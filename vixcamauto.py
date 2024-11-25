from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pymediainfo import MediaInfo
from openpyxl import Workbook, load_workbook
import requests
import time
import logging
import os

# 터미널에 로그출력
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 메인페이지
def setup_driver():
    options = Options()
    options.add_experimental_option("detach", True)
    options.add_argument("--headless") 
    service = Service()
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(3)
    return driver

# 로그인
def login(driver, username, password):
    driver.get('http://172.16.15.108/')
    action = ActionChains(driver)
    driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div/form/p[1]/input').click()
    action.send_keys(username).key_down(Keys.TAB).send_keys(password).pause(1).key_down(Keys.ENTER).perform()
    action.reset_actions()
    time.sleep(5)
    logger.info('로그인 성공')

# 설정로그인 + 비디오 스트림 3, 4 h.264로 설정
def loginop(driver, username, password):
    driver.get('http://' + input('ip주소를 입력하세요'))
    action = ActionChains(driver)
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div/form/p[3]/div/ins').click()
    driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div/form/p[1]/input').click()
    username = input('ID를 입력하세요')
    password = input('패스워드를 입력하세요')
    action.send_keys(username).key_down(Keys.TAB).send_keys(password).pause(1).key_down(Keys.ENTER).perform()
    action.reset_actions()
    time.sleep(5)
    logger.info('설정로그인 성공')
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[2]/a').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[2]/ul/li[2]/a').click()
    time.sleep(3)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[1]/select/option[1]').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[2]/select/option[1]').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div[1]/a[3]').click()
    logger.info('비디오 스트림 3,4 H264 코덱으로 설정 완료')

# 탭 호출 > 라이브
def tab_l(driver):
    driver.find_element(By.XPATH, '/html/body/div[1]/div/button').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[1]/a[1]').click()
    logger.info('라이브 화면으로 이동')

# 탭 호출 > 플레이백
def tab_p(driver):
    driver.find_element(By.XPATH, '/html/body/div[1]/div/button').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[1]/a[2]').click()
    logger.info('플레이백 화면으로 이동')

# 탭 호출 > 설정
def tab_o(driver):
    driver.find_element(By.XPATH, '/html/body/div[1]/div/button').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[1]/a[3]').click()
    logger.info('설정 화면으로 이동')

# 탭 호출 > 로그아웃
def tab_bye(driver):
    driver.find_element(By.XPATH, '/html/body/div[1]/div/button').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[1]/a[4]').click()
    logger.info('로그아웃')

# 설정 > 비디오 스트림 진입
def op_stream(driver):
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[2]/a').click()
    streamgo = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/ul/li[2]/ul/li[2]/a')))
    streamgo.click()
    logger.info('비디오 스트림 진입 완료')

# 플레이백 > 비디오 스트림 진입
def op_stream_pb(driver):
    driver.find_element(By.XPATH, '/html/body/div[1]/div/button').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div[1]/a[3]/i').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[2]/a').click()
    streamgo = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/ul/li[2]/ul/li[2]/a')))
    streamgo.click()
    logger.info('비디오 스트림 진입 완료')


# 저장소 겹쳐쓰기 설정
def reckeep(driver):
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[4]/a').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/form/div/div[2]/div[1]/div[1]/label/div/ins').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('저장소 겹쳐쓰기 설정 완료')

# 이벤트 녹화 사용설정-비디오 스트림1
def recon(driver):
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/a').click()
    driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/ul/li[2]/a').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '//*[@id="tab-record"]/div[1]/label/div').click()
    driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('비디오 스트림1로 이벤트 녹화 사용설정 완료')

# 이벤트 녹화 사용설정-비디오 스트림2
def recon2(driver):
    driver.find_element(By.XPATH, '/html/body/div[1]/div/button').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div[1]/a[3]').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/a').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/ul/li[2]/a').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[1]/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[1]/div[2]/select/option[2]').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('비디오 스트림2로 이벤트 녹화 사용설정 완료')

# 이벤트 녹화 사용설정-비디오 스트림3
def recon3(driver):
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/a').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/ul/li[2]/a').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[1]/div[2]/select').click()
    time.sleep(3)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[1]/div[2]/select/option[3]').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('비디오 스트림3로 이벤트 녹화 사용설정 완료')

# 이벤트 녹화 사용설정-비디오 스트림4
def recon4(driver):
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/a').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/ul/li[2]/a').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[1]/div[2]/select').click()
    time.sleep(3)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[1]/div[2]/select/option[4]').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('비디오 스트림4로 이벤트 녹화 사용설정 완료')

# 이벤트 녹화 설정(타이머)
def rectimer(driver):
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[2]/div[1]/ul/li[5]/a'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/ul/li[1]/a'))).click()
    time.sleep(5)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[1]/ul/li[9]/a'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tab-timer"]/div[1]/label/div'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[3]/select'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[3]/select/option[1]'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[4]/select'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[4]/select/option[2]'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button'))).click()
    logger.info('타이머 트리거 설정 완료 - 1분!')
    time.sleep(5)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/ul/li[3]/a'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[1]/div/div[2]/input[1]'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div[2]/form/div[2]/div[2]/div/div[1]/select'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div[2]/form/div[2]/div[2]/div/div[1]/select/option[46]'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="modal_rule"]/div/div/div[2]/form/div[3]/div[2]/div[8]/div/label/div'))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div[3]/button[1]'))).click()
    logger.info('이벤트 규칙 설정완료 - 타이머')

# 이벤트 영상 추출하기
def videosave(driver, videoname):
    driver.find_element(By.XPATH, '//*[@id="sidebar-shortcuts-large"]/a[2]').click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="1"]/td[2]'))).click()
    video_element = driver.find_element(By.TAG_NAME, 'video')
    video_url = video_element.get_attribute('src')
    response = requests.get(video_url, timeout=60)
    directory = "D:/testvideo"
    os.makedirs(directory, exist_ok=True)
    file_name = f"{videoname}.mp4"
    file_path = os.path.join(directory, file_name)
    file = open(file_path, 'wb') 
    file.write(response.content)
    logger.info("비디오 파일이 %s으로 저장되었습니다.", file_name)


#엑셀 저장
def excelsave(file_path, sheet_name, cell, log_message):
 
    if not os.path.exists(file_path):
        if not os.path.exists(os.path.dirname(file_path)):
         os.makedirs(os.path.dirname(file_path))
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws[cell] = log_message
        wb.save(file_path)
        print(f"지정한 디렉토리에 결과값만 작성한 파일을 생성합니다. TC양식에 붙여넣으세요.")
    else:
        print(f"지정한 디렉토리에 TC 파일이 이미 존재합니다. 데이터를 업데이트합니다.")
        workbook = load_workbook(file_path)  
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook.create_sheet(sheet_name)  
        ws[cell] = log_message
        workbook.save(file_path) 
        workbook.close() 


# 이벤트 영상 조건 확인하기
def info(videoname):
    directory = os.path.join("D:", "testvideo")
    file_path = os.path.join(directory, f"{videoname}")
    file_path = os.path.join("D:", "testvideo", f"{videoname}.mp4")
    media_info = MediaInfo.parse(file_path)
    result = []
    for track in media_info.tracks:
        if track.track_type == 'Video':
            result.append(f"코덱, {track.codec_id}")
            result.append(f"코덱 프로파일, {track.format_profile}")
            result.append(f"영상길이, {track.duration}")
            result.append(f"fps, {track.frame_rate}")
            result.append(f"width, {track.width}")
            result.append(f"height, {track.height}")
            result.append(f"비트레이트 모드, {track.bit_rate_mode}")
            result.append(f"비트레이트, {track.bit_rate}")
    return "\n".join(result)

# cama-1 압축방식 High
def cama1(driver):
    videoname = 'cama1'
    # 압축 방식 H.264 High
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select/option[1]').click()
    # 해상도 2160
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select/option[1]').click()
    # 프레임레이트 30
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[1]').click()
    # GOP크기 60
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[1]').click()
    # 비트레이트 제어 CBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select/option[2]').click()
    # 비트레이트 6000
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[7]/select/option[101]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-1 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E3", log_message = info(videoname))

# cama-2 압축방식 Main
def cama2(driver):
    videoname = 'cama2'
    # 압축 방식 H.264 Main
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select/option[2]').click()
    # 해상도 2160
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select/option[1]').click()
    # 프레임레이트 30
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[1]').click()
    # GOP크기 60
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[1]').click()
    # 비트레이트 제어 CBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select/option[2]').click()
    # 비트레이트 6000
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[7]/select/option[101]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-2 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="D3", log_message = info(videoname))

# cama-3 압축방식 Smart
def cama3(driver):
    videoname = 'cama3'
    # 압축 방식 H.264+ Smart
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select/option[3]').click()
    # 해상도 2160
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select/option[1]').click()
    # 프레임레이트 30
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[1]').click()
    # GOP크기 60
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[1]').click()
    # 비트레이트 제어 CBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select/option[2]').click()
    # 비트레이트 6000
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[7]/select/option[101]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-3 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="F3", log_message = info(videoname))

# cama-4 해상도 1728, 프레임 1
def cama4(driver):
    videoname = 'cama4'
    # 압축 방식 H.264 High
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select/option[1]').click()
    # 해상도 1728
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select/option[2]').click()
    # 프레임레이트 1
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[30]').click()
    # GOP크기 60
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[1]').click()
    # 비트레이트 제어 CBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select/option[2]').click()
    # 비트레이트 6000
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[7]/select/option[101]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-4 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="G3", log_message = info(videoname))

# cama-5 해상도 1440, 프레임 15, 비트레이트12000
def cama5(driver):
    videoname = 'cama5'
    # 압축 방식 H.264 High
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select/option[1]').click()
    # 해상도 1440
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select/option[3]').click()
    # 프레임레이트 1
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[15]').click()
    # GOP크기 60
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[1]').click()
    # 비트레이트 제어 CBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select/option[2]').click()
    # 비트레이트 12000
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[7]/select/option[1]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-5 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="H3", log_message = info(videoname))

# cama-6 비트레이트 제어 VBR
def cama6(driver):
    videoname = 'cama6'
    # 압축 방식 H.264 High
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[1]/select/option[1]').click()
    # 해상도 1440
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[2]/select/option[3]').click()
    # 프레임레이트 1
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[15]').click()
    # GOP크기 60
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[3]/select/option[1]').click()
    # 비트레이트 제어 VBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[1]/div/div[2]/div/div[6]/select/option[2]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-6 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="I3", log_message = info(videoname))

# cama-7 ************************비디오스트림2 필수지정!!*******************************
def cama7(driver):
    videoname = 'cama7'
    # 압축 방식 H.264 High
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select/option[1]').click()
    # 해상도 1080
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select/option[1]').click()
    # 프레임레이트 30
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select/option[1]').click()
    # GOP크기 60
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select/option[1]').click()
    # 비트레이트 제어 CBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[6]/select/option[2]').click()
    # 비트레이트 12000
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[7]/select/option[1]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-7 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="J3", log_message = info(videoname))

# cama-8
def cama8(driver):
    videoname = 'cama8'
    # 압축 방식 H.264 Main
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select/option[2]').click()
    # 해상도 896
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select/option[2]').click()
    # 프레임레이트 1
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select/option[30]').click()
    # GOP크기 60
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select/option[1]').click()
    # 비트레이트 제어 CBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[6]/select/option[2]').click()
    # 비트레이트 6000
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[7]/select/option[61]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-8 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="K3", log_message = info(videoname))

#  # cama-9
#  def cama9(driver):
#      videoname = 'cama9'
#      # 압축 방식 MJPEG
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select/option[6]').click()
#      # 해상도 720
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select/option[3]').click()
#      # 프레임레이트 30
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select/option[1]').click()
#      # 품질 99
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[5]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[5]/select/option[1]').click()
#      # 저장
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
#      logger.info('CAMa-9 TC Precondition 설정 완료')
#      time.sleep(150)
#      videosave(driver, videoname)

# cama-10
def cama10(driver):
    videoname = 'cama10'
    # 압축 방식 H.264 Main
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select/option[2]').click()
    # 해상도 432
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select/option[4]').click()
    # 프레임레이트 30
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select/option[1]').click()
    # GOP크기 60
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select/option[1]').click()
    # 비트레이트 제어 CBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[6]/select/option[2]').click()
    # 비트레이트 4500
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[7]/select/option[36]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-10 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="L3", log_message = info(videoname))

# cama-11
def cama11(driver):
    videoname = 'cama11'
    # 압축 방식 H.264 High
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select/option[1]').click()
    # 해상도 392
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select/option[5]').click()
    # 프레임레이트 20
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select/option[11]').click()
    # GOP크기 60
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select/option[1]').click()
    # 비트레이트 제어 VBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[6]/select/option[1]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-11 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="M3", log_message = info(videoname))

# cama-12
def cama12(driver):
    videoname = 'cama12'
    # 압축 방식 H.264 High
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[1]/select/option[1]').click()
    # 해상도 360
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[2]/select/option[5]').click()
    # 프레임레이트 1
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select/option[30]').click()
    # GOP크기 60
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[3]/select/option[1]').click()
    # 비트레이트 제어 VBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[2]/div/div[2]/div/div[6]/select/option[1]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-12 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="N3", log_message = info(videoname))

# cama-13 *********************비디오스트림3 필수 지정!!!!!!**************************
def cama13(driver):
    videoname = 'cama13'
    # 압축 방식 H.264 High
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[1]/select/option[1]').click()
    # 해상도 432
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[2]/select/option[1]').click()
    # 프레임레이트 30
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[3]/select/option[1]').click()
    # GOP크기 10
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[4]/select/option[246]').click()
    # 비트레이트 제어 CBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[6]/select/option[2]').click()
    # 비트레이트 8000
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[7]/select/option[1]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-13 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="O3", log_message = info(videoname))

# cama-14
def cama14(driver):
    videoname = 'cama14'
    # 압축 방식 H.264 Main
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[1]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[1]/select/option[2]').click()
    # 해상도 392
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[2]/select/option[2]').click()
    # 프레임레이트 15
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[3]/select/option[16]').click()
    # GOP크기 10
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[4]/select/option[246]').click()
    # 비트레이트 제어 VBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[6]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[6]/select/option[1]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-14 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="P3", log_message = info(videoname))

# cama-15
#  def cama15(driver):
#      videoname = 'cama15'
#      # 압축 방식 MJPEG
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[1]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[1]/select/option[6]').click()
#      # 해상도 360
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[2]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[2]/select/option[3]').click()
#      # 프레임레이트 1
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[3]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[3]/select/option[30]').click()
#      # 품질 10
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[5]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[3]/div/div[2]/div/div[5]/select/option[10]').click
#      # 저장
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
#      logger.info('CAMa-15 TC Precondition 설정 완료')
#      time.sleep(150)
#      videosave(driver, videoname)

# cama-16 *****************비디오스트림4 필수 지정!!!!!!!!!!!!!!!!**************************************
def cama16(driver):
    videoname = 'cama16'
    # 압축 방식 H.264 High
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[2]/select/option[1]').click()
    # 해상도 720
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[3]/select/option[1]').click()
    # 프레임레이트 30
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[4]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[4]/select/option[1]').click()
    # GOP크기 10
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[5]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[5]/select/option[246]').click()
    # 비트레이트 제어 VBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[7]/select/option[1]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-16 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="Q3", log_message = info(videoname))

# cama-17
def cama17(driver):
    videoname = 'cama17'
    # 압축 방식 H.264 Main
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[2]/select/option[2]').click()
    # 해상도 432
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[3]/select/option[2]').click()
    # 프레임레이트 30
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[4]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[4]/select/option[1]').click()
    # GOP크기 100
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[5]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[5]/select/option[156]').click()
    # 비트레이트 제어 VBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[7]/select/option[1]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-17 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="R3", log_message = info(videoname))
# cama-18
def cama18(driver):
    videoname = 'cama18'
    # 압축 방식 H.264 Main
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[2]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[2]/select/option[2]').click()
    # 해상도 392
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[3]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[3]/select/option[3]').click()
    # 프레임레이트 15
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[4]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[4]/select/option[16]').click()
    # GOP크기 10
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[5]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[5]/select/option[246]').click()
    # 비트레이트 제어 CBR
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[7]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[7]/select/option[2]').click()
    # 비트레이트 4000
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[8]/select').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[8]/select/option[41]').click()
    # 저장
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
    logger.info('CAMa-18 TC Precondition 설정 완료')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="S3", log_message = info(videoname))

# cama-19
#  def cama19(driver):
#      videoname = 'cama19'
#      # 압축 방식 MJPEG
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[2]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[2]/select/option[3]').click()
#      # 해상도 360
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[3]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[3]/select/option[3]').click()
#      # 프레임레이트 15
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[4]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[4]/select/option[16]').click()
#      # 품질 30
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[6]/select').click()
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[3]/div/form/div[4]/div/div[2]/div/div[6]/select/option[8]').click()
#      # 저장
#      driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button').click()
#      logger.info('CAMa-19 TC Precondition 설정 완료')
#      time.sleep(150)
#      videosave(driver, videoname)    
# # 테스트 진행 전 무조건 저장소 초기화, 시스템 초기화 후 비밀번호 초기 설정까지 완료하세요
def main():
    driver = setup_driver()
    loginop(driver, 'admin', 'pass0001!')
    time.sleep(5)
    recon(driver)
    time.sleep(5)
    reckeep(driver)
    time.sleep(5)
    rectimer(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama1(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama2(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama3(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama4(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama5(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama6(driver)
    time.sleep(5)
    recon2(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama7(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama8(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama10(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama11(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama12(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    recon3(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama13(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama14(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    recon4(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama16(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama17(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama18(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)

if __name__ == "__main__":
    main()
